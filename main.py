import json
import os
import re
import sys
import unicodedata
import numpy
from googleapiclient.discovery import build
from pprint import pprint
import datetime
import pandas
import openpyxl
import sqlite3
import const

pandas.options.display.max_rows = None
pandas.options.display.max_columns = None
pandas.options.display.width = 6000
pandas.options.display.max_colwidth = 6000
pandas.options.display.colheader_justify = 'left'

API_KEY = os.getenv('YTV3_API_KEY', default='')
print(os.environ['DEBUG'], type(os.environ['DEBUG']))

if API_KEY == '':
    print('No API Key.')
    exit(-1)
youtube = build(
    'youtube',
    'v3',
    developerKey=API_KEY
)

workbookName = 'save.xlsx'
process_list = const.playlists()

# process_list = [
#     ['PL59O0JbKsFLp-RYKpMhN_pooj_bJckLL7', 'アップアップガールズ(仮)']
# ]
count = 0


def trim_title(text, artist_name):
    title_regex_one = r"[\(（]([a-zA-Z\s\[\]\'\"”””“\.…,\/　!！=。@’[°C]・:〜\-])*[\)）]"
    title_regex_two = r"[\s|-][Promotion|Music|Video].*[\s|\-]*|Edit|カバー|[\s|-]*YouTube[\s|\-]*|ver|【セルフカヴァー】" \
                      r"|Promotion|Music|Video"
    title_regex_three = r"[\[\(（]([a-zA-Z\s”\[［\]］\.\/’\'\"&。:〜”“0-9\-=\?!×#~,♂　@（）])*[\]\)）]?"
    title_regex_four = r'※.*'  # r"ショート|Short|short|Version|Ver.|Ver|バージョン|Dance|ダンス|リリック.*|"
    if artist_name == '鞘師里保':
        return re.sub(r'\(.*\)', '', text)
    if artist_name == 'COVERS - One on One -':
        return text
    if artist_name == 'アップアップガールズ(仮)':
        tmp = re.sub(r'【MUSIC VIDEO】|\([A-z\s\[\]\-！!].*\)|ミュージック|イメージ|ビデオ|\[.*?\]', '', text,
                     re.IGNORECASE)
        return re.sub(r'UP|GIRLS|kakko|KARI|MV|MUSIC|VIDEO|（MV）', '', tmp, re.IGNORECASE | re.MULTILINE)
    if artist_name == 'シャ乱Q':
        return re.sub(r'[v|(].*', '', unicodedata.normalize('NFKC', text))
    if artist_name == 'つんく♂' or artist_name == '℃-ute' or artist_name == 'Berryz工房' or artist_name == 'Buono!' \
            or artist_name == 'PINK CRES.' or artist_name == 'こぶしファクトリー' or artist_name == 'ブラザーズ5' \
            or artist_name == '真野恵里菜' or artist_name == '鈴木愛理':
        return re.sub(r'\(.*', '', unicodedata.normalize('NFKC', text))
    if artist_name == 'Bitter & Sweet':
        return re.sub(r'\(.*?\)', '', unicodedata.normalize('NFKC', text))
    if artist_name == 'HANGRY&ANGRY':
        return unicodedata.normalize('NFKC', text).replace('VIDEO CLIP', '')
    if artist_name == 'Juice=Juice' or artist_name == 'アンジュルム' or artist_name == '高橋愛・田中れいな・夏焼雅' \
            or artist_name == 'BEYOOOOONDS' or artist_name == 'モーニング娘。' or artist_name == 'OCHA NORMA':
        return re.sub(r'\(.*?\)|\[.*?]|MV|Promotion|Edit', '', unicodedata.normalize('NFKC', text))
    if artist_name == 'LoVendoЯ':
        return re.sub(r'\(.*|\[.*', '', unicodedata.normalize('NFKC', text))
    if artist_name == '小片リサ':
        return re.sub(r'\-.*', '', unicodedata.normalize('NFKC', text))

    text = str(text).replace('(仮)', '@kari@')
    text = str(text).replace('（仮）', '@kari@')
    text = str(text).replace('℃-ute', '@cute@')
    text = str(text).replace('°C-ute', '@cute@')
    return_code = re.sub(title_regex_four, '',
                         re.sub(title_regex_three, '',
                                re.sub(title_regex_two, '',
                                       re.sub(title_regex_one, ' ', unicodedata.normalize('NFKC', text),
                                              re.IGNORECASE), re.IGNORECASE | re.MULTILINE), re.IGNORECASE),
                         re.IGNORECASE)
    return return_code.replace('@kari@', '(仮)').replace('@cute@', '℃-ute')


def get_view_count_and_data(Id, artist_name):
    global count
    video_info_raw = youtube.videos().list(part='statistics,snippet',
                                           fields='items(snippet/title,statistics/viewCount)',
                                           id=Id).execute()['items']
    video_info = json.loads(unicodedata.normalize('NFKC', json.dumps(video_info_raw)))

    if not video_info:
        return None
    count += 1
    print(count, end='\t')
    print('https://youtu.be/' + Id, end='\t')
    print(trim_title(video_info[0]['snippet']['title'], artist_name))
    sys.stdout.flush()
    sys.stderr.flush()
    # if 'maxres' in video_info[0]['snippet']['thumbnails']:
    #     thumb = video_info[0]['snippet']['thumbnails']['maxres']
    #     # print('maxres')
    # else:
    #     thumb = video_info[0]['snippet']['thumbnails']['high']

    data = [unicodedata.normalize('NFKC', video_info[0]['snippet']['title']),
            video_info[0]['statistics']['viewCount'],
            'https://youtu.be/' + Id]
    return data


def extract_playlist(playlist_key):
    nextPageToken = ''
    return_videoId = []
    while True:
        playlist = youtube.playlistItems().list(part='snippet',
                                                fields='items/snippet/resourceId/videoId,nextPageToken',
                                                playlistId=playlist_key,
                                                maxResults=50,
                                                pageToken=nextPageToken).execute()

        # pprint(playlist['items'])
        for Id in playlist['items']:
            return_videoId.append(Id['snippet']['resourceId']['videoId'])
        if 'nextPageToken' not in playlist:
            break
        nextPageToken = str(playlist['nextPageToken'])
        # break
        # print('\t' + nextPageToken)
        # time.sleep(3)
    return return_videoId


def process_channel(artistName, playlistKey):
    channelInfo = youtube.channels().list(
        part='snippet,brandingSettings,statistics',
        fields='items(snippet/thumbnails/high,'
               'brandingSettings/image,'
               'statistics(subscriberCount,videoCount,viewCount))',
        id=youtube.playlistItems().list(part='snippet',
                                        fields='items/snippet/channelId',
                                        playlistId=playlistKey,
                                        maxResults=1).execute()['items'][0]['snippet']['channelId']).execute()
    pprint(channelInfo)

    # exit(0)

    today = str(datetime.date.today())
    if os.path.isfile(workbookName):
        workbook = openpyxl.load_workbook(workbookName)
    else:
        workbook = openpyxl.Workbook()

    if artistName not in workbook.sheetnames:
        dataframe = pandas.DataFrame([0])
    else:
        dataframe = pandas.read_excel('save.xlsx', sheet_name=artistName, index_col=0)

    if 'タイトル' not in dataframe.columns:
        dataframe['タイトル'] = ''

    if today not in dataframe.columns:
        dataframe[today] = 0

    for Id in extract_playlist(playlist_key=playlistKey):
        # print(Id)
        data = get_view_count_and_data(Id=Id, artist_name=artistName)
        # pprint(data)
        if data is None:
            continue
        title, viewCount, url = data
        if url not in dataframe.index.tolist():
            dataframe.loc[url] = 0

        dataframe.at[url, 'タイトル'] = trim_title(title, artistName)

        dataframe.at[url, today] = int(viewCount)
    if 0 in dataframe.columns:
        dataframe.drop(0, axis=0, inplace=True)
        dataframe.drop(0, axis=1, inplace=True)

    if not os.path.isfile(workbookName):
        workbook = openpyxl.Workbook()
        if 'Sheet' in workbook.sheetnames and len(workbook.sheetnames) != 1:
            workbook.remove(workbook['Sheet'])
        workbook.save(workbookName)

        dataframe.replace(0, numpy.NaN)

    connector = sqlite3.connect(os.path.join('save.sqlite'))
    dataframe.to_sql(artistName, connector, if_exists='replace')
    connector.close()

    with pandas.ExcelWriter(workbookName, mode='a', if_sheet_exists='replace') as writer:
        dataframe.to_excel(writer, sheet_name=artistName)

    os.makedirs('tsvs', exist_ok=True)
    dataframe.to_csv(os.path.join(os.getcwd(), 'tsvs', artistName + '.tsv'), sep='\t')

    workbook = openpyxl.load_workbook(workbookName)
    if 'Sheet' in workbook.sheetnames and len(workbook.sheetnames) != 1:
        workbook.remove(workbook['Sheet'])
    for sheet in workbook.worksheets:
        sheet.freeze_panes = 'C2'
        rows = sheet[1]
        for row in rows:
            sheet.column_dimensions[row.column_letter].width = 14
        sheet.column_dimensions['C'].width = 12
        sheet.column_dimensions['A'].width = 45
        sheet.column_dimensions['B'].width = 60
    workbook.save(workbookName)

    # print(pandas.read_excel('save.xlsx', sheet_name=artistName, index_col=0))


for processes in process_list:
    count = 0
    process_channel(artistName=processes[1], playlistKey=processes[0])

os.makedirs(os.path.join(os.getcwd(), 'tsvs'), exist_ok=True)
with open(os.path.join(os.getcwd(), 'tsvs', 'group_list.tsv'), mode='w',encoding='utf8') as f:
    f.writelines('\n'.join([i for _, i in process_list]))
